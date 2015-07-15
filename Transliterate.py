import os
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
from transliterate import translit

def progress_bar(curr_val,end_val,label, bar_length=20):
        percent = float(curr_val) / end_val
        hashes = '#' * int(round(percent * bar_length))
        spaces = ' ' * (bar_length - len(hashes))
        sys.stdout.write("\r{0}: [{1}] {2}%".format(label,hashes + spaces, int(round(percent * 100))))
        sys.stdout.flush()


def process_line(line):

    newVal = translit(line,'bg2',reversed=True)

    if line.isupper():
        return newVal.upper()

    return newVal




def translate_file(filename):
    print("Converting " + filename)
    wb = load_workbook(filename=filename, read_only=True)
    toWb = Workbook(write_only=True)

    sheet_names = wb.get_sheet_names()

    for sheet in sheet_names:
        count = 0
        ws = wb[sheet]
        max_rows = ws.get_highest_row()
        toWs = toWb.create_sheet()
        toWs.title = ws.title

        for row in ws.rows:
            newRow = []
            for cell in row:
                if isinstance(cell.value, str):
                    newRow.append(process_line(cell.value))
                else:
                    newRow.append(cell.value)

            toWs.append(newRow)
            if count % 100 == 0:
                # sys.stdout.write("\rProcessing " + sheet + ": row:" + str(count))
                progress_bar(count,max_rows,filename + ":" + sheet)

            count += 1

        print("")

    fp = os.path.splitext(filename)
    fn = fp[0] + "_converted" + fp[1]
    print("Saving file: " + fn)
    toWb.save(fn)

if len(sys.argv) == 1:
    print("You must provide at least one file to translate...")
else:
    for i in range(1, len(sys.argv)):
        f = sys.argv[i]
        translate_file(f)

    print("Processing complete")



