from openpyxl import load_workbook
import sys
from CategoriesAccumulatorClass import CategoriesAccumulator
from DegreeAccumulator import DegreeAccumulator
from ExperienceAccumulator import ExperienceAccumulator


def progress_bar(curr_val, end_val, label, bar_length=20):
    percent = float(curr_val) / end_val
    hashes = '#' * int(round(percent * bar_length))
    spaces = ' ' * (bar_length - len(hashes))
    sys.stdout.write("\r{0}: [{1}] {2}%".format(label, hashes + spaces, int(round(percent * 100))))
    sys.stdout.flush()


def addEmUp(filename):
    print("Adding up " + filename + "...")

    degreesAcc = DegreeAccumulator(4)
    categoriesAcc = CategoriesAccumulator(11)
    experienceAcc = ExperienceAccumulator(16)

    wb = load_workbook(filename=filename, read_only=True)
    sheet_names = wb.get_sheet_names()
    for sheet in sheet_names:
        count = 0
        ws = wb[sheet]
        max_rows = ws.get_highest_row()

        prevId = ws['B2'].value

        degreesAcc.newSheet(ws.title)
        categoriesAcc.newSheet(ws.title)
        experienceAcc.newSheet(ws.title)

        for row in ws.rows:

            # skip the first row
            if row[0] == ws['A1']:
                continue

            idSchool = row[1].value
            if idSchool == prevId:
                degreesAcc.accumulateRow(row)
                categoriesAcc.accumulateRow(row)
                experienceAcc.accumulateRow(row)

            else:
                degreesAcc.appendRow(row[0].value, prevId)
                categoriesAcc.appendRow(row[0].value, prevId)
                experienceAcc.appendRow(row[0].value, prevId)
                prevId = idSchool

            count += 1
            if count % 100 == 0:
                # sys.stdout.write("\rProcessing " + sheet + ": row:" + str(count))
                progress_bar(count, max_rows, filename + ":" + sheet)


        # end for loop newRows = createDegreeRows(row[0], prevId, degrees, map_degree)
        degreesAcc.appendRow(row[0].value, idSchool)
        categoriesAcc.appendRow(row[0].value, idSchool)
        experienceAcc.appendRow(row[0].value, idSchool)

    print("")
    degreesAcc.writeFile(filename, "degrees")
    categoriesAcc.writeFile(filename, "categories")
    experienceAcc.writeFile(filename, "experience")


# __________ MAIN _______________
if len(sys.argv) == 1:
    print("You must provide at least one file to translate...")
else:
    for i in range(1, len(sys.argv)):
        f = sys.argv[i]
        addEmUp(f)

    print("Processing complete")
